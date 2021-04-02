from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Max, Min
from django.shortcuts import render, redirect
from django import forms
import openpyxl
import colorsys
import json

from airpollution.models import Pollutant, Country, PollutantEntry
from airpollution.helpers import get_headers_and_units, XLHEADERS


class ExcelUploadForm(forms.Form):
    year = forms.CharField(max_length=4)
    file = forms.FileField()

# Create your views here.


def airpollution(request):
    ctx = {
        'app_name': request.resolver_match.app_name,
        'pollutant_list': [p.name for p in Pollutant.objects.all()]
    }
    if request.method == 'GET':
        pass
    elif request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            year = form.cleaned_data['year']
            file = form.cleaned_data['file']
            wb = openpyxl.load_workbook(filename=file, read_only=False)
            tab_names = wb.get_sheet_names()
            for tab_name in tab_names:
                ws = wb[tab_name]
                pollutant_name = tab_name.split('_')[0].strip()
                pollutant = Pollutant.objects.get_or_create(
                    name=pollutant_name)

                if pollutant[0].limit_value is None:
                    limit_value = int(ws['A'][2].value.split()[-2])
                    pollutant[0].limit_value = limit_value
                    pollutant[0].save()
                headers_row, headers, units = get_headers_and_units(ws)

                # Save all entrties to database
                to_insert = []
                for i, row in enumerate(ws.rows):
                    if i <= headers_row:  # Skip to actual entries
                        continue

                    country = row[headers[XLHEADERS.COUNTRY]].value
                    if country is None:
                        break

                    if len(country) > 2:
                        country_object = Country.objects.filter(
                            name=country).first()
                    else:
                        country_object = Country.objects.get(pk=country)

                    city = row[headers[XLHEADERS.CITY]].value
                    station_name = row[headers[XLHEADERS.CITY]].value
                    station_area = row[headers[XLHEADERS.AREA]].value

                    data = {
                        'pollutant': pollutant[0],
                        'country': country_object,
                        'year': year,
                        'city': city if city else '',
                        'station_code': row[headers[XLHEADERS.STATION_EOI_CODE]].value,
                        'station_name': station_name if station_name else '',
                        'pollution_level': row[headers[XLHEADERS.AIR_POLLUTION_LEVEL]].value,
                        'units': units,
                        'station_type': row[headers[XLHEADERS.TYPE]].value,
                        'station_area': station_area if station_area else '',
                        'longitude': row[headers[XLHEADERS.LONGITUDE]].value,
                        'latitude': row[headers[XLHEADERS.LATITUDE]].value,
                        'altitude': row[headers[XLHEADERS.ALTITUDE]].value,
                    }

                    to_insert.append(PollutantEntry(**data))
                    print(to_insert)
                    PollutantEntry.objects.filter(
                        year=year, pollutant=pollutant[0]).delete()
                    PollutantEntry.objects.bulk_create(to_insert)
        ctx['message_sucess'] = 'File uploaded successfully!!'
    else:
        return HttpResponse('This view only handles GET and POST request')
    return render(request, 'airpollution/welcome.html', ctx)


def airpollution_table_data(request):
    table_data = {}

    pollutant_list = [pollutant for pollutant in Pollutant.objects.all()]
    country_list = [country for country in Country.objects.all()]
    for pollutant in pollutant_list:
        table_data[pollutant.name] = {}

        for i, country in enumerate(country_list):
            total = PollutantEntry.objects.aggregate(total=Sum('pollution_level',
                                                               filter=Q(pollutant=pollutant, country=country)))['total']

            minimun = PollutantEntry.objects.aggregate(min=Min('pollution_level',
                                                               filter=Q(pollutant=pollutant, country=country)))['min']

            maximun = PollutantEntry.objects.aggregate(max=Max('pollution_level',
                                                               filter=Q(pollutant=pollutant, country=country)))['max']
            count = PollutantEntry.objects.filter(
                pollutant=pollutant, country=country).count()
            units = PollutantEntry.objects.filter(
                pollutant=pollutant, country=country).first()
            units = units.units if units else ''
            if total is not None and count:
                table_data[pollutant.name][country.iso_code] = {
                    'avg': total / count,
                    'min': minimun,
                    'max': maximun,
                    'limit': pollutant.limit_value,
                    'units': units}

    return JsonResponse(table_data)


def airpollution_visual_data1(request):
    visuals_data = {'Pollutions Levels by Pollutant by Country': {
        'chart_type': 'chart1',
        'labels': [],
        'datasets': [
            {
                'label': 'limit',
                'backgroundColor': '#338DD8',
                'stack': 'limit',
                'data': [],
            }
        ]
    }}
    pollutant_list = [pollutant for pollutant in Pollutant.objects.all()]
    country_list = [country for country in Country.objects.all()]
    visuals_data['Pollutions Levels by Pollutant by Country']['datasets'] += \
        [{'label': c.name, 'backgroundColor': c.color,
            'hidden': 'true', 'data': []} for c in country_list]

    for pollutant in pollutant_list:
        visuals_data['Pollutions Levels by Pollutant by Country']['labels'].append(
            pollutant.name)
        visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][0]['data'].append(
            pollutant.limit_value)
        for i, country in enumerate(country_list):
            total = PollutantEntry.objects.aggregate(total=Sum('pollution_level',
                                                               filter=Q(pollutant=pollutant, country=country)))['total']

            count = PollutantEntry.objects.filter(
                pollutant=pollutant, country=country).count()

            if total is not None and count:
                visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][i +
                                                                                      1]['data'].append(round(total/count, 2))
            else:
                visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][i +
                                                                                      1]['data'].append(-1)

    return JsonResponse(visuals_data)


def airpollution_visual_data2(request):
    pollutant_name = request.GET.get('pollutant', 'PM10')
    pollutant = Pollutant.objects.get(name=pollutant_name)
    sumary_type = request.GET.get('sumary_type', 'max')
    if sumary_type == 'avg':
        name_prefix = 'Average'
    elif sumary_type == 'min':
        name_prefix = 'Mininum'
    else:
        name_prefix = 'Maxinum'
    all_years = [pe['year'] for pe in PollutantEntry.objects.order_by(
        'year').values('year').distinct()]
    all_countries = list(Country.objects.all())
    all_pollutions = [p.name for p in Pollutant.objects.all()]

    visuals_data = {
        'name': f'{name_prefix} pollution level by country over time',
        'labels': all_years,
        'datasets':  [
            {
                'label': 'limit',
                'backgroundColor': '#338DD8',
                'borderColor': '#338DD8',
                'data': [pollutant.limit_value] * len(all_years),
                'fill': False
            }
        ]
    }

    for country in all_countries:
        country_data = {
            'label': country.name,
            'backgroundColor': country.color,
            'borderColor': country.color,
            'data': [],
            'fill': False,
            'hidden': True
        }

        visuals_data['datasets'].append(country_data)

        for year in all_years:
            f = Q(pollutant=pollutant, year=year, country=country)

            if sumary_type == 'avg':
                country_tot = PollutantEntry.objects.aggregate(s=Sum('pollution_level', filter=f))['s']
                country_count = PollutantEntry.objects.filter(f).count()
                country_data['data'].append(country_tot / country_count if country_count else 0)
            elif sumary_type == 'min':
                country_min = PollutantEntry.objects.aggregate(s=Min('pollution_level', filter=f))['s']
                country_data['data'].append(country_min if country_min else 0)
            else: # by default -> max
                country_max = PollutantEntry.objects.aggregate(s=Min('pollution_level', filter=f))['s']
                country_data['data'].append(country_max if country_max else 0)



    return JsonResponse(visuals_data)


# def airpollution_backup(request):
#     if request.method == 'GET':
#         table_data = {}
#         visuals_data = {'Pollutions Levels by Pollutant by Country': {
#             'chart_type': 'chart1',
#             'labels': [],
#             'datasets': [
#                 {
#                     'label': 'limit',
#                     'backgroundColor': '#338DD8',
#                     'stack': 'limit',
#                     'data': [],
#                 }
#             ]
#         }}
#         pollutant_list = [pollutant for pollutant in Pollutant.objects.all()]
#         country_list = [country for country in Country.objects.all()]
#         visuals_data['Pollutions Levels by Pollutant by Country']['datasets'] += \
#             [{'label': c.name, 'backgroundColor': c.color,
#                 'hidden': 'true', 'data': []} for c in country_list]

#         for pollutant in pollutant_list:
#             table_data[pollutant.name] = {}
#             visuals_data['Pollutions Levels by Pollutant by Country']['labels'].append(
#                 pollutant.name)
#             visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][0]['data'].append(
#                 pollutant.limit_value)
#             for i, country in enumerate(country_list):
#                 total = PollutantEntry.objects.aggregate(total=Sum('pollution_level',
#                                                                    filter=Q(pollutant=pollutant, country=country)))['total']

#                 minimun = PollutantEntry.objects.aggregate(min=Min('pollution_level',
#                                                                    filter=Q(pollutant=pollutant, country=country)))['min']

#                 maximun = PollutantEntry.objects.aggregate(max=Max('pollution_level',
#                                                                    filter=Q(pollutant=pollutant, country=country)))['max']
#                 count = PollutantEntry.objects.filter(
#                     pollutant=pollutant, country=country).count()
#                 units = PollutantEntry.objects.filter(
#                     pollutant=pollutant, country=country).first()
#                 units = units.units if units else ''
#                 if total is not None and count:
#                     table_data[pollutant.name][country.iso_code] = {
#                         'avg': total / count, 'min': minimun, 'max': maximun, 'limit': pollutant.limit_value, 'units': units}
#                     visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][i +
#                                                                                           1]['data'].append(round(total/count, 2))
#                 else:
#                     visuals_data['Pollutions Levels by Pollutant by Country']['datasets'][i +
#                                                                                           1]['data'].append(-1)
#         # Post procces visual data
#         for pollutant_data in visuals_data.values():

#             pollutant_data['labels'] = json.dumps(pollutant_data['labels'])
#             pollutant_data['datasets'] = json.dumps(pollutant_data['datasets'])

#         ctx = {
#             'app_name': request.resolver_match.app_name,
#             'data': table_data,
#             'visuals_data': visuals_data
#         }

#     elif request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             year = form.cleaned_data['year']
#             file = form.cleaned_data['file']
#             wb = openpyxl.load_workbook(filename=file, read_only=False)
#             tab_names = wb.get_sheet_names()
#             for tab_name in tab_names:
#                 ws = wb[tab_name]
#                 pollutant_name = tab_name.split('_')[0].strip()
#                 pollutant = Pollutant.objects.get_or_create(
#                     name=pollutant_name)

#                 if pollutant[0].limit_value is None:
#                     limit_value = int(ws['A'][2].value.split()[-2])
#                     pollutant[0].limit_value = limit_value
#                     pollutant[0].save()
#                 headers_row, headers, units = get_headers_and_units(ws)

#                 # Save all entrties to database
#                 to_insert = []
#                 for i, row in enumerate(ws.rows):
#                     if i <= headers_row:  # Skip to actual entries
#                         continue

#                     country = row[headers[XLHEADERS.COUNTRY]].value
#                     if country is None:
#                         break

#                     if len(country) > 2:
#                         country_object = Country.objects.filter(
#                             name=country).first()
#                     else:
#                         country_object = Country.objects.get(pk=country)

#                     city = row[headers[XLHEADERS.CITY]].value
#                     station_name = row[headers[XLHEADERS.CITY]].value
#                     station_area = row[headers[XLHEADERS.AREA]].value

#                     data = {
#                         'pollutant': pollutant[0],
#                         'country': country_object,
#                         'year': year,
#                         'city': city if city else '',
#                         'station_code': row[headers[XLHEADERS.STATION_EOI_CODE]].value,
#                         'station_name': station_name if station_name else '',
#                         'pollution_level': row[headers[XLHEADERS.AIR_POLLUTION_LEVEL]].value,
#                         'units': units,
#                         'station_type': row[headers[XLHEADERS.TYPE]].value,
#                         'station_area': station_area if station_area else '',
#                         'longitude': row[headers[XLHEADERS.LONGITUDE]].value,
#                         'latitude': row[headers[XLHEADERS.LATITUDE]].value,
#                         'altitude': row[headers[XLHEADERS.ALTITUDE]].value,
#                     }

#                     to_insert.append(PollutantEntry(**data))
#                     print(to_insert)
#                     PollutantEntry.objects.filter(
#                         year=year, pollutant=pollutant[0]).delete()
#                     PollutantEntry.objects.bulk_create(to_insert)
#         ctx = {
#             'app_name': request.resolver_match.app_name,
#             'message_sucess': 'File uploaded successfully!!'
#         }

#     else:  # Request method not on POST
#         return HttpResponse('This view only handles GET and POST request')
#     return render(request, 'airpollution/welcome.html', ctx)


def temp_country_creator(request):
    countries = {
        'Albania': ['AL', '#f68a8a'],
        'Andorra': ['AD', '#2019c0'],
        'Austria': ['AT', '#a81b1a'],
        'Belgium': ['BE', '#808080'],
        'Bosnia and Herzegovina': ['BA', '#ffd208'],
        'Bulgaria': ['BG', '#468658'],
        'Croatia': ['HR', '#21248a'],
        'Cyprus': ['CY', '#1fdb94'],
        'Czech Republic': ['CZ', '#1f48db'],
        'Denmark': ['DK', '#b68008'],
        'Estonia': ['EE', '#a056ae'],
        'Finland': ['FI', '#b68000'],
        'France': ['FR', '#dfdf6a'],
        'Germany': ['DE', '#e0be1d'],
        'Greece': ["GR", '#0b66aa'],
        'Hungary': ['HU', '#295934'],
        'Iceland': ['IS', '#2933d9'],
        'Ireland': ['IE', '#00d641'],
        'Italy': ['IT', '#338DD8'],
        'Kosovo under UNSRC 1244/99': ['XK', '#33d857'],
        'Latvia': ['LV', '#431f58'],
        'Lithuania': ['LT', '#3698ae'],
        'Luxembourg': ['LU', '#0f924c'],
        'Malta': ['MT', '#a19087'],
        'Montenegro': ['ME', '#a97641'],
        'Netherlands': ['NL', '#ec3000'],
        'Norway': ['NO', '#ec8f00'],
        'Poland': ['PL', '#616f55'],
        'Portugal': ['PT', '#0aa76c'],
        'Romania': ['RO', '#05464b'],
        'Serbia': ['RS', '#17355e'],
        'Slovakia': ['SK', '#c528b7'],
        'Slovenia': ['SI', '#45093f'],
        'Span': ['ES', '#fd0066'],
        'Sweden': ['SE', '#5c9a2c'],
        'Switzerland': ['CH', '#f00'],
        'The former Yugoslav Republic of Macedonia': ['MK', '#6d6f6b'],
        'Turkey': ['TR', '#00F9FF'],
        'United Kingdom': ['GB', '#e3ff00']
    }

    to_insert = []
    for country_name, data in countries.items():
        to_insert.append(
            Country(iso_code=data[0], name=country_name, color=data[1]))

    if request.GET.get('update', '') == 'true':
        Country.objects.bulk_update(to_insert, ['color'])
    else:
        Country.objects.bulk_create(to_insert)

    return redirect('airpollution:airpollution')


def temp_add_colors_to_pollutants(request):
    pollutants = ['PM2.5', 'PM10', 'NO2', 'O3', 'BaP', 'SO2']
    colors = ['#09363c', '#2b2726', '#e76b27', '#00601a', '#596907', '#b5c7c5']

    to_insert = [Pollutant(name=pollutant, color=colors[i])
                 for i, pollutant in enumerate(pollutants)]
    # print(to_insert)
    Pollutant.objects.bulk_update(to_insert, ['color'])
    return redirect('airpollution:airpollution')
